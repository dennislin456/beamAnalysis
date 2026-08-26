import os
import json
import pandas as pd
import numpy as np
import pyqtgraph as pg
import pyqtgraph.exporters as pg_export
from PIL import Image
import openpyxl
from openpyxl.chart import LineChart, Reference

from PyQt5.QtWidgets import (QWidget, QHBoxLayout, QVBoxLayout, QPushButton, 
                             QLabel, QFileDialog, QMessageBox, QFrame, 
                             QSplitter, QSizePolicy, QRadioButton, QButtonGroup, 
                             QScrollArea, QCheckBox, QComboBox, QApplication)
from PyQt5.QtCore import Qt

# 💡 匯入共用的元件 (稍後會在 shared_components.py 提供)
from shared_components import (
    NoWheelSpinBox, NoWheelDoubleSpinBox, CrossProfileViewerWindow,
    apply_readable_plot_theme, LevelAlignedHistogramLUTItem,
)

class BaslerTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # --- Basler 專屬變數初始化 ---
        self.save_dir_path = ""
        self.basler_img_path = ""
        self.basler_param_path = ""
        self.basler_matrix = None
        self.basler_circle_item = None
        self.basler_center_spot = None
        self.basler_center = None 
        self.basler_fixed_cross_items = []
        self.is_updating_basler_ui = False

        self.basler_click_points = []
        self.basler_measure_items = []
        self.basler_heatmap_cross_point = None
        self.basler_heatmap_cross_items = []
        self.basler_cross_profile_win = None

        # 建立 UI
        self.setup_ui()

    def setup_ui(self):
        btn_style_default = """
            QPushButton { font-size: 13px; font-weight: bold; background-color: #f0f0f0; border: 1px solid #cccccc; border-radius: 5px; padding: 6px 12px; }
            QPushButton:hover { background-color: #e0e0e0; border-color: #b0b0b0; }
            QPushButton:pressed { background-color: #d0d0d0; }
        """
        btn_style_folder = """
            QPushButton { font-size: 13px; font-weight: bold; color: white; background-color: #EF6C00; border: none; border-radius: 5px; padding: 6px 12px; }
            QPushButton:hover { background-color: #F57C00; }
            QPushButton:pressed { background-color: #E65100; }
        """
        btn_style_cross = """
            QPushButton { font-size: 13px; font-weight: bold; color: white; background-color: #00897B; border: none; border-radius: 5px; padding: 6px 14px; }
            QPushButton:hover { background-color: #009688; }
            QPushButton:pressed { background-color: #00695C; }
            QPushButton:disabled { background-color: #E0E0E0; color: #A0A0A0; }
        """
        btn_style_primary = """
            QPushButton { font-size: 13px; font-weight: bold; color: white; background-color: #2E7D32; border: none; border-radius: 5px; padding: 8px 12px; }
            QPushButton:hover { background-color: #388E3C; }
            QPushButton:pressed { background-color: #1B5E20; }
        """
        btn_style_export = """
            QPushButton { font-size: 13px; font-weight: bold; color: white; background-color: #0288D1; border: none; border-radius: 5px; padding: 8px 16px; }
            QPushButton:hover { background-color: #039BE5; }
            QPushButton:pressed { background-color: #01579B; }
            QPushButton:disabled { background-color: #B0BEC5; }
        """

        basler_layout = QHBoxLayout(self)
        basler_layout.setContentsMargins(6, 6, 6, 6)
        
        splitter = QSplitter(Qt.Horizontal)
        splitter.setStyleSheet("QSplitter::handle { background-color: #dcdcdc; width: 4px; }")
        basler_layout.addWidget(splitter)
        
        # --- 左側主容器 ---
        basler_left_container = QWidget()
        basler_left_container.setMinimumWidth(320)
        basler_left_container.setMaximumWidth(360)
        basler_left_outer_layout = QVBoxLayout(basler_left_container)
        basler_left_outer_layout.setContentsMargins(0, 0, 0, 0)
        basler_left_outer_layout.setSpacing(0)

        # 1. 上方固定面板
        basler_top_fixed_widget = QWidget()
        basler_top_fixed_layout = QVBoxLayout(basler_top_fixed_widget)
        basler_top_fixed_layout.setContentsMargins(12, 12, 12, 12)
        basler_top_fixed_layout.setSpacing(6)

        lbl_basler_mode_title = QLabel("選擇工作模式：")
        lbl_basler_mode_title.setStyleSheet("font-weight: bold; color: #333333; margin-top: 2px;")
        basler_top_fixed_layout.addWidget(lbl_basler_mode_title)
        
        self.combo_basler_mode = QComboBox()
        self.combo_basler_mode.addItem("單獨匯入畫圖", "single")
        self.combo_basler_mode.addItem("匯入圖像/數據 + 載入參數檔重繪", "param")
        self.combo_basler_mode.currentIndexChanged.connect(self.on_basler_mode_changed)
        basler_top_fixed_layout.addWidget(self.combo_basler_mode)

        basler_top_fixed_layout.addWidget(self._create_hline())

        self.btn_load_bmp = QPushButton("I. 匯入 BMP/PNG/CSV 檔")
        self.btn_load_bmp.setStyleSheet(btn_style_default)
        self.btn_load_bmp.clicked.connect(self.load_basler_bmp)
        basler_top_fixed_layout.addWidget(self.btn_load_bmp)
        
        self.lbl_bmp_path = QLabel("未選擇圖像/數據檔")
        self.lbl_bmp_path.setStyleSheet("color: #757575; font-size: 11px;")
        self.lbl_bmp_path.setWordWrap(True)
        basler_top_fixed_layout.addWidget(self.lbl_bmp_path)

        self.btn_load_basler_param = QPushButton("II. 匯入 Basler JSON 參數檔")
        self.btn_load_basler_param.setStyleSheet(btn_style_default)
        self.btn_load_basler_param.clicked.connect(self.load_basler_param_file)
        self.btn_load_basler_param.setVisible(False)
        basler_top_fixed_layout.addWidget(self.btn_load_basler_param)

        self.lbl_basler_param_path = QLabel("未選擇參數檔")
        self.lbl_basler_param_path.setStyleSheet("color: #757575; font-size: 11px;")
        self.lbl_basler_param_path.setWordWrap(True)
        self.lbl_basler_param_path.setVisible(False)
        basler_top_fixed_layout.addWidget(self.lbl_basler_param_path)
        
        basler_top_fixed_layout.addWidget(self._create_hline())

        self.btn_basler_calc = QPushButton("單獨讀取並畫圖")
        self.btn_basler_calc.setStyleSheet(btn_style_primary)
        self.btn_basler_calc.setMinimumHeight(38)
        self.btn_basler_calc.clicked.connect(self.process_basler_data)
        basler_top_fixed_layout.addWidget(self.btn_basler_calc)

        basler_top_fixed_layout.addWidget(self._create_hline())

        lbl_basler_status_title = QLabel("計算數據與狀態")
        lbl_basler_status_title.setStyleSheet("font-weight: bold; font-size: 12px; color: #37474F;")
        basler_top_fixed_layout.addWidget(lbl_basler_status_title)

        self.lbl_basler_status = QLabel("狀態: 等待匯入檔案")
        self.lbl_basler_status.setStyleSheet("color: #1565C0; font-weight: bold; font-size: 12px;")
        basler_top_fixed_layout.addWidget(self.lbl_basler_status)

        self.lbl_basler_distance = QLabel("位置差距: ΔX: --, ΔY: -- | 總距離: -- px")
        self.lbl_basler_distance.setStyleSheet("font-weight: bold; color: #37474F; font-size: 12px;")
        basler_top_fixed_layout.addWidget(self.lbl_basler_distance)

        self.lbl_basler_real_distance = QLabel("實際差距 (* 3.45): ΔX: -- μm, ΔY: -- μm | 總距離: -- μm")
        self.lbl_basler_real_distance.setStyleSheet("font-weight: bold; color: #2E7D32; font-size: 12px;")
        basler_top_fixed_layout.addWidget(self.lbl_basler_real_distance)

        basler_left_outer_layout.addWidget(basler_top_fixed_widget)

        # 2. 下方捲動面板
        basler_scroll_panel = QScrollArea()
        basler_scroll_panel.setWidgetResizable(True)

        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(12, 12, 12, 12)
        left_layout.setSpacing(10)

        left_layout.addWidget(self._create_hline())

        lbl_center_mode = QLabel("中心定位模式：")
        lbl_center_mode.setStyleSheet("font-weight: bold;")
        left_layout.addWidget(lbl_center_mode)

        self.chk_basler_enable_spot = QCheckBox("啟用光斑抓取與標記")
        self.chk_basler_enable_spot.setChecked(True)
        self.chk_basler_enable_spot.setStyleSheet("font-weight: bold; color: #2E7D32;")
        self.chk_basler_enable_spot.toggled.connect(self.recalculate_basler_spot)
        left_layout.addWidget(self.chk_basler_enable_spot)

        self.basler_center_group = QButtonGroup(self)
        self.radio_auto_center = QRadioButton("自動抓取 (最高值區域幾何中心)")
        self.radio_auto_center.setChecked(True)
        self.radio_auto_center.toggled.connect(self.recalculate_basler_spot)
        self.basler_center_group.addButton(self.radio_auto_center)
        left_layout.addWidget(self.radio_auto_center)

        self.radio_centroid_center = QRadioButton("自動抓取 (質心中心)")
        self.radio_centroid_center.toggled.connect(self.recalculate_basler_spot)
        self.basler_center_group.addButton(self.radio_centroid_center)
        left_layout.addWidget(self.radio_centroid_center)

        self.radio_thresh_geom_center = QRadioButton("自動抓取 (門檻區域幾何中心)")
        self.radio_thresh_geom_center.toggled.connect(self.recalculate_basler_spot)
        self.basler_center_group.addButton(self.radio_thresh_geom_center)
        left_layout.addWidget(self.radio_thresh_geom_center)

        self.radio_manual_center = QRadioButton("手動抓取 (點擊影像中心)")
        self.radio_manual_center.toggled.connect(self.recalculate_basler_spot)
        self.basler_center_group.addButton(self.radio_manual_center)
        left_layout.addWidget(self.radio_manual_center)
        
        left_layout.addWidget(self._create_hline())

        self.chk_basler_heatmap_cross = QCheckBox("啟用點擊定格十字標記")
        self.chk_basler_heatmap_cross.setChecked(True)
        self.chk_basler_heatmap_cross.setStyleSheet("font-weight: bold; color: #00897B;")
        self.chk_basler_heatmap_cross.toggled.connect(self.on_basler_heatmap_cross_toggled)
        left_layout.addWidget(self.chk_basler_heatmap_cross)

        self.chk_basler_measure_cross = QCheckBox("啟用兩點距離量測十字標記")
        self.chk_basler_measure_cross.setChecked(False)
        self.chk_basler_measure_cross.setStyleSheet("font-weight: bold; color: #E65100;")
        self.chk_basler_measure_cross.toggled.connect(self.on_basler_measure_cross_toggled)
        left_layout.addWidget(self.chk_basler_measure_cross)

        left_layout.addWidget(self._create_hline())

        lbl_shape_title = QLabel("抓取光斑形狀選項：")
        lbl_shape_title.setStyleSheet("font-weight: bold; color: #1565C0;")
        left_layout.addWidget(lbl_shape_title)

        self.basler_shape_group = QButtonGroup(self)

        layout_shape_radios = QHBoxLayout()
        self.radio_shape_circle = QRadioButton("正圓 (Circle)")
        self.radio_shape_circle.setChecked(True)
        self.radio_shape_circle.toggled.connect(self.on_shape_type_changed)
        
        self.radio_shape_ellipse = QRadioButton("橢圓 (Ellipse)")
        self.radio_shape_ellipse.toggled.connect(self.on_shape_type_changed)

        self.basler_shape_group.addButton(self.radio_shape_circle)
        self.basler_shape_group.addButton(self.radio_shape_ellipse)

        layout_shape_radios.addWidget(self.radio_shape_circle)
        layout_shape_radios.addWidget(self.radio_shape_ellipse)
        left_layout.addLayout(layout_shape_radios)
        
        self.container_circle_spin = QWidget()
        layout_circle_spin = QHBoxLayout(self.container_circle_spin)
        layout_circle_spin.setContentsMargins(0, 0, 0, 0)
        lbl_dia = QLabel("光斑直徑 Circle Size (px):")
        self.spin_circle_diameter = NoWheelSpinBox()
        self.spin_circle_diameter.setRange(2, 5000)
        self.spin_circle_diameter.setValue(100)
        self.spin_circle_diameter.setSingleStep(5)
        self.spin_circle_diameter.valueChanged.connect(self.update_basler_circle)
        layout_circle_spin.addWidget(lbl_dia)
        layout_circle_spin.addWidget(self.spin_circle_diameter)
        left_layout.addWidget(self.container_circle_spin)

        self.container_ellipse_spin = QWidget()
        layout_ellipse_spin = QVBoxLayout(self.container_ellipse_spin)
        layout_ellipse_spin.setContentsMargins(0, 0, 0, 0)

        layout_ellipse_x = QHBoxLayout()
        lbl_ellipse_x = QLabel("X 軸直徑 Wx (px):")
        self.spin_ellipse_wx = NoWheelSpinBox()
        self.spin_ellipse_wx.setRange(2, 5000)
        self.spin_ellipse_wx.setValue(100)
        self.spin_ellipse_wx.setSingleStep(5)
        self.spin_ellipse_wx.valueChanged.connect(self.update_basler_circle)
        layout_ellipse_x.addWidget(lbl_ellipse_x)
        layout_ellipse_x.addWidget(self.spin_ellipse_wx)

        layout_ellipse_y = QHBoxLayout()
        lbl_ellipse_y = QLabel("Y 軸直徑 Wy (px):")
        self.spin_ellipse_wy = NoWheelSpinBox()
        self.spin_ellipse_wy.setRange(2, 5000)
        self.spin_ellipse_wy.setValue(100)
        self.spin_ellipse_wy.setSingleStep(5)
        self.spin_ellipse_wy.valueChanged.connect(self.update_basler_circle)
        layout_ellipse_y.addWidget(lbl_ellipse_y)
        layout_ellipse_y.addWidget(self.spin_ellipse_wy)

        layout_ellipse_spin.addLayout(layout_ellipse_x)
        layout_ellipse_spin.addLayout(layout_ellipse_y)
        self.container_ellipse_spin.setVisible(False)
        left_layout.addWidget(self.container_ellipse_spin)

        self.chk_use_threshold = QCheckBox("使用門檻計算光斑寬度")
        self.chk_use_threshold.setChecked(True)
        self.chk_use_threshold.setStyleSheet("font-weight: bold; color: #C62828;")
        self.chk_use_threshold.toggled.connect(self.on_threshold_toggled)
        left_layout.addWidget(self.chk_use_threshold)

        layout_thresh = QHBoxLayout()
        self.lbl_thresh_spin = QLabel("門檻比例 Threshold (%):")
        self.spin_thresh_percent = NoWheelDoubleSpinBox()
        self.spin_thresh_percent.setRange(0.1, 100.0)
        self.spin_thresh_percent.setValue(13.5)
        self.spin_thresh_percent.setSingleStep(1.0)
        self.spin_thresh_percent.setDecimals(1)
        self.spin_thresh_percent.valueChanged.connect(self.recalculate_basler_spot)
        layout_thresh.addWidget(self.lbl_thresh_spin)
        layout_thresh.addWidget(self.spin_thresh_percent)
        left_layout.addLayout(layout_thresh)

        left_layout.addWidget(self._create_hline())

        lbl_b_measure_title = QLabel("距離量測與分色十字工具")
        lbl_b_measure_title.setStyleSheet("font-weight: bold; font-size: 13px; color: #C62828;")
        left_layout.addWidget(lbl_b_measure_title)
        
        self.lbl_basler_cursor = QLabel("目前滑鼠： X: -- , Y: --")
        self.lbl_basler_cursor.setStyleSheet("color: #616161;")
        left_layout.addWidget(self.lbl_basler_cursor)
        
        self.lbl_basler_measure_points = QLabel("點擊點 1 (黃): --\n點擊點 2 (青藍): --")
        left_layout.addWidget(self.lbl_basler_measure_points)
        
        layout_b_cross_ctrl = QHBoxLayout()
        lbl_b_cross_size = QLabel("標記十字大小 (px):")
        self.spin_basler_cross_size = NoWheelSpinBox()
        self.spin_basler_cross_size.setRange(10, 200)
        self.spin_basler_cross_size.setValue(40)
        self.spin_basler_cross_size.valueChanged.connect(self.redraw_basler_measure_crosses)
        layout_b_cross_ctrl.addWidget(lbl_b_cross_size)
        layout_b_cross_ctrl.addWidget(self.spin_basler_cross_size)
        left_layout.addLayout(layout_b_cross_ctrl)

        self.btn_clear_basler_measure = QPushButton("清除量測點與十字")
        self.btn_clear_basler_measure.setStyleSheet(btn_style_default)
        self.btn_clear_basler_measure.clicked.connect(self.clear_basler_measure_points)
        left_layout.addWidget(self.btn_clear_basler_measure)

        left_layout.addWidget(self._create_hline())

        self.lbl_b_center = QLabel("中心座標: (X: --, Y: --)")
        self.lbl_b_peak = QLabel("最大強度 (Peak): --")
        self.lbl_b_thresh_val = QLabel("計算門檻值: --")
        self.lbl_b_x_width = QLabel("X 軸寬度 (@門檻): -- px (-- μm)")
        self.lbl_b_y_width = QLabel("Y 軸寬度 (@門檻): -- px (-- μm)")
        self.lbl_b_dia_px = QLabel("外框尺寸: -- px")
        self.lbl_b_dia_um = QLabel("實際尺寸: -- μm")
        self.lbl_b_radius_um = QLabel("實際半徑: -- μm")
        self.lbl_b_area_um = QLabel("實際面積: -- μm²")
        self.lbl_b_sum_intensity = QLabel("總光強度: --")
        self.lbl_b_mean_intensity = QLabel("平均強度: --")
        
        left_layout.addWidget(self.lbl_b_center)
        left_layout.addWidget(self.lbl_b_peak)
        left_layout.addWidget(self.lbl_b_thresh_val)
        left_layout.addWidget(self.lbl_b_x_width)
        left_layout.addWidget(self.lbl_b_y_width)
        left_layout.addWidget(self._create_hline())
        left_layout.addWidget(self.lbl_b_dia_px)
        left_layout.addWidget(self.lbl_b_dia_um)
        left_layout.addWidget(self.lbl_b_radius_um)
        left_layout.addWidget(self.lbl_b_area_um)
        left_layout.addWidget(self.lbl_b_sum_intensity)
        left_layout.addWidget(self.lbl_b_mean_intensity)
        
        left_layout.addStretch()
        basler_scroll_panel.setWidget(left_panel)
        basler_left_outer_layout.addWidget(basler_scroll_panel)
        
        # --- 右側主繪圖區 ---
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(6, 6, 6, 6)
        
        top_bar = QHBoxLayout()
        top_bar.setContentsMargins(0, 0, 0, 6)

        self.btn_basler_select_dir = QPushButton("選擇儲存資料夾")
        self.btn_basler_select_dir.setStyleSheet(btn_style_folder)
        self.btn_basler_select_dir.setMinimumHeight(38)
        self.btn_basler_select_dir.clicked.connect(self.select_save_directory)
        top_bar.addWidget(self.btn_basler_select_dir)

        self.lbl_basler_top_file1_path = QLabel("")
        self.lbl_basler_top_file1_path.setStyleSheet("color: #333333; background-color: #f5f5f5; border: 1px solid #d0d0d0; border-radius: 4px; padding: 6px 10px; font-size: 12px;")
        self.lbl_basler_top_file1_path.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.lbl_basler_top_file1_path.setMinimumHeight(38)
        top_bar.addWidget(self.lbl_basler_top_file1_path)

        self.btn_basler_view_cross = QPushButton("查看十字波形視窗")
        self.btn_basler_view_cross.setStyleSheet(btn_style_cross)
        self.btn_basler_view_cross.setMinimumHeight(38)
        self.btn_basler_view_cross.setEnabled(False)
        self.btn_basler_view_cross.clicked.connect(self.show_basler_cross_profile_window)
        top_bar.addWidget(self.btn_basler_view_cross)

        self.btn_export_basler = QPushButton("匯出 Basler 結果 (JSON/Excel/PNG)")
        self.btn_export_basler.setStyleSheet(btn_style_export)
        self.btn_export_basler.setMinimumHeight(38)
        self.btn_export_basler.setEnabled(False)
        self.btn_export_basler.clicked.connect(self.export_basler_results)
        top_bar.addWidget(self.btn_export_basler)
        right_layout.addLayout(top_bar)
        
        right_splitter = QSplitter(Qt.Vertical)
        
        self.win_basler = pg.GraphicsLayoutWidget()
        
        colors = [(0, 0, 255), (0, 255, 255), (0, 255, 0), (255, 255, 0), (255, 0, 0)]
        pos = np.linspace(0.0, 1.0, len(colors))
        jet_map = pg.ColorMap(pos, colors)
        
        self.plot_basler = self.win_basler.addPlot(row=0, col=0, title='Basler Beam Spot Analysis')
        self.plot_basler.getViewBox().invertY(False)
        self.plot_basler.setAspectLocked(True)
        self.plot_basler.setLabel('bottom', 'X Pixels')
        self.plot_basler.setLabel('left', 'Y Pixels')
        
        self.basler_image_item = pg.ImageItem()
        self.plot_basler.addItem(self.basler_image_item)

        self.basler_v_line = pg.InfiniteLine(angle=90, movable=False, pen=pg.mkPen('#37474F', width=1, style=Qt.DashLine))
        self.basler_h_line = pg.InfiniteLine(angle=0, movable=False, pen=pg.mkPen('#37474F', width=1, style=Qt.DashLine))
        self.plot_basler.addItem(self.basler_v_line, ignoreBounds=True)
        self.plot_basler.addItem(self.basler_h_line, ignoreBounds=True)
        self.basler_v_line.hide()
        self.basler_h_line.hide()
        
        self.basler_hist = LevelAlignedHistogramLUTItem()
        self.basler_hist.setImageItem(self.basler_image_item)
        self.basler_hist.gradient.setColorMap(jet_map)
        self.basler_hist.sigLevelsChanged.connect(self.on_basler_colorbar_changed)
        self.win_basler.addItem(self.basler_hist, row=0, col=1)

        apply_readable_plot_theme(self.win_basler, [self.plot_basler])
        
        self.plot_basler.scene().sigMouseMoved.connect(self.on_basler_mouse_moved)
        self.plot_basler.scene().sigMouseClicked.connect(self.on_basler_scene_clicked)
        right_splitter.addWidget(self.win_basler)

        bottom_horizontal_splitter = QSplitter(Qt.Horizontal)

        scroll_x_column = QScrollArea()
        scroll_x_column.setWidgetResizable(True)
        scroll_x_column.setStyleSheet("QScrollArea { border: 1px solid #90A4AE; background-color: #E8EEF2; }")

        self.win_basler_x_col = pg.GraphicsLayoutWidget()
        self.win_basler_x_col.setMinimumHeight(480)

        self.plot_basler_x_profile = self.win_basler_x_col.addPlot(row=0, col=0, title="X-Axis Intensity Profile (Linear Scale)")
        self.plot_basler_x_profile.setLabel('bottom', 'X Position (px)')
        self.plot_basler_x_profile.setLabel('left', 'Intensity')
        self.plot_basler_x_profile.showGrid(x=True, y=True, alpha=0.3)

        self.plot_basler_x_log_profile = self.win_basler_x_col.addPlot(row=1, col=0, title="X-Axis Log Intensity Profile Log10(Intensity)")
        self.plot_basler_x_log_profile.setLabel('bottom', 'X Position (px)')
        self.plot_basler_x_log_profile.setLabel('left', 'Log10(Intensity)')
        self.plot_basler_x_log_profile.showGrid(x=True, y=True, alpha=0.3)

        apply_readable_plot_theme(
            self.win_basler_x_col,
            [self.plot_basler_x_profile, self.plot_basler_x_log_profile],
        )
        scroll_x_column.setWidget(self.win_basler_x_col)

        scroll_y_column = QScrollArea()
        scroll_y_column.setWidgetResizable(True)
        scroll_y_column.setStyleSheet("QScrollArea { border: 1px solid #90A4AE; background-color: #E8EEF2; }")

        self.win_basler_y_col = pg.GraphicsLayoutWidget()
        self.win_basler_y_col.setMinimumHeight(480)

        self.plot_basler_y_profile = self.win_basler_y_col.addPlot(row=0, col=0, title="Y-Axis Intensity Profile (Linear Scale)")
        self.plot_basler_y_profile.setLabel('bottom', 'Y Position (px)')
        self.plot_basler_y_profile.setLabel('left', 'Intensity')
        self.plot_basler_y_profile.showGrid(x=True, y=True, alpha=0.3)

        self.plot_basler_y_log_profile = self.win_basler_y_col.addPlot(row=1, col=0, title="Y-Axis Log Intensity Profile Log10(Intensity)")
        self.plot_basler_y_log_profile.setLabel('bottom', 'X Position (px)')
        self.plot_basler_y_log_profile.setLabel('left', 'Log10(Intensity)')
        self.plot_basler_y_log_profile.showGrid(x=True, y=True, alpha=0.3)

        apply_readable_plot_theme(
            self.win_basler_y_col,
            [self.plot_basler_y_profile, self.plot_basler_y_log_profile],
        )
        scroll_y_column.setWidget(self.win_basler_y_col)

        bottom_horizontal_splitter.addWidget(scroll_x_column)
        bottom_horizontal_splitter.addWidget(scroll_y_column)
        bottom_horizontal_splitter.setSizes([500, 500])

        right_splitter.addWidget(bottom_horizontal_splitter)
        right_splitter.setStretchFactor(0, 2)
        right_splitter.setStretchFactor(1, 1)
        right_splitter.setSizes([600, 300])

        right_layout.addWidget(right_splitter)
        
        splitter.addWidget(basler_left_container)
        splitter.addWidget(right_panel)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        splitter.setSizes([340, 1060])

    # =========================================================================
    # 事件與邏輯層
    # =========================================================================
    def _create_hline(self):
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        line.setStyleSheet("color: #e0e0e0; margin-top: 2px; margin-bottom: 2px;")
        return line

    def _compute_auto_spot_center(self, matrix, mode, use_threshold=False, thresh_percent=50.0):
        matrix = np.asarray(matrix, dtype=np.float64)
        h, w = matrix.shape
        peak_val = float(np.max(matrix)) if matrix.size > 0 else 0.0

        if mode == "peak_geom":
            max_y_indices, max_x_indices = np.where(matrix == peak_val)
            if len(max_x_indices) > 0:
                return int(round(np.mean(max_x_indices))), int(round(np.mean(max_y_indices)))
            peak_idx = np.unravel_index(np.argmax(matrix, axis=None), matrix.shape)
            return int(peak_idx[1]), int(peak_idx[0])

        thresh_val = peak_val * (thresh_percent / 100.0) if use_threshold else peak_val * 0.5
        mask = matrix >= thresh_val
        if not np.any(mask):
            peak_idx = np.unravel_index(np.argmax(matrix, axis=None), matrix.shape)
            return int(peak_idx[1]), int(peak_idx[0])

        ys, xs = np.where(mask)
        if mode == "thresh_geom":
            return int(round(np.mean(xs))), int(round(np.mean(ys)))

        weights = matrix[mask]
        wsum = float(np.sum(weights))
        if wsum <= 0:
            return int(round(np.mean(xs))), int(round(np.mean(ys)))
        cx = int(round(np.sum(xs * weights) / wsum))
        cy = int(round(np.sum(ys * weights) / wsum))
        cx = max(0, min(w - 1, cx))
        cy = max(0, min(h - 1, cy))
        return cx, cy

    def on_basler_mode_changed(self):
        mode_data = self.combo_basler_mode.currentData()
        is_param_mode = (mode_data == "param")
        self.btn_load_basler_param.setVisible(is_param_mode)
        self.lbl_basler_param_path.setVisible(is_param_mode)
        
        if is_param_mode:
            self.btn_basler_calc.setText("載入參數並重繪圖表")
        else:
            self.btn_basler_calc.setText("單獨讀取並畫圖")

    def on_basler_heatmap_cross_toggled(self, checked):
        if checked:
            self.chk_basler_measure_cross.blockSignals(True)
            self.chk_basler_measure_cross.setChecked(False)
            self.chk_basler_measure_cross.blockSignals(False)
            self.clear_basler_measure_items_only()
        self.redraw_basler_heatmap_cross_item()

    def on_basler_measure_cross_toggled(self, checked):
        if checked:
            self.chk_basler_heatmap_cross.blockSignals(True)
            self.chk_basler_heatmap_cross.setChecked(False)
            self.chk_basler_heatmap_cross.blockSignals(False)
            for item in self.basler_heatmap_cross_items:
                self.plot_basler.removeItem(item)
            self.basler_heatmap_cross_items.clear()
        self.redraw_basler_measure_crosses()

    def select_save_directory(self):
        dir_path = QFileDialog.getExistingDirectory(self, "選擇儲存資料夾", "")
        if dir_path:
            self.save_dir_path = dir_path
            self.lbl_basler_top_file1_path.setText(f"{dir_path}")

    def show_basler_cross_profile_window(self):
        if self.basler_matrix is not None:
            if self.basler_cross_profile_win is not None:
                try:
                    self.basler_cross_profile_win.close()
                    self.basler_cross_profile_win.deleteLater()
                except Exception:
                    pass
                self.basler_cross_profile_win = None

            self.basler_cross_profile_win = CrossProfileViewerWindow("Basler", self)
            self.basler_cross_profile_win.show()
            
            cx, cy = self.basler_heatmap_cross_point if self.basler_heatmap_cross_point is not None else (self.basler_matrix.shape[1]//2, self.basler_matrix.shape[0]//2)
            min_v, max_v = self.basler_hist.getLevels()
            self.basler_cross_profile_win.update_profiles(self.basler_matrix, cx, cy, y_range=(min_v, max_v))

    def load_basler_bmp(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "選擇圖像或 CSV 數據檔", "", 
            "All Supported Files (*.csv *.bmp *.png *.jpg *.jpeg);;CSV Files (*.csv);;Image Files (*.bmp *.png *.jpg *.jpeg)"
        )
        if path:
            self.basler_img_path = path
            self.lbl_bmp_path.setText(os.path.basename(path))
            self.lbl_bmp_path.setStyleSheet("color: #212121; font-size: 11px;")
            if not self.save_dir_path:
                self.save_dir_path = os.path.dirname(path)
            if self.combo_basler_mode.currentData() == "param":
                QApplication.processEvents()
                self.load_basler_param_file()

    def load_basler_param_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "選擇 Basler JSON 參數檔", "", "JSON Files (*.json)")
        if path:
            self.basler_param_path = path
            self.lbl_basler_param_path.setText(os.path.basename(path))
            self.lbl_basler_param_path.setStyleSheet("color: #212121; font-size: 11px;")

    def process_basler_data(self):
        if not self.basler_img_path:
            QMessageBox.warning(self, "警告", "請先匯入 BMP/PNG/CSV 檔案！")
            return
            
        if self.combo_basler_mode.currentData() == "param" and not self.basler_param_path:
            QMessageBox.warning(self, "警告", "請選取要帶入的 JSON 參數檔！")
            return

        try:
            self.clear_basler_measure_points()
            self.lbl_basler_status.setText("狀態: 正在讀取與分析檔案...")
            self.lbl_basler_status.setStyleSheet("color: #F57C00; font-weight: bold;")
            QApplication.processEvents()

            ext = os.path.splitext(self.basler_img_path)[1].lower()
            if ext == '.csv':
                df = pd.read_csv(self.basler_img_path, header=None, skiprows=25)
                arr = df.dropna(how='all').astype(float).values
            else:
                img = Image.open(self.basler_img_path).convert('L')
                arr = np.array(img, dtype=float)

            self.basler_matrix = np.flipud(arr)
            
            self.basler_image_item.setImage(self.basler_matrix.T)
            min_v, max_v = float(np.min(self.basler_matrix)), float(np.max(self.basler_matrix))
            self.basler_hist.setHistogramRange(min_v, max_v, padding=0)
            self.basler_hist.setLevels(min_v, max_v)
            
            self.btn_export_basler.setEnabled(True)
            self.btn_basler_view_cross.setEnabled(True)

            if self.combo_basler_mode.currentData() == "param":
                self.apply_basler_param_file()
            else:
                self.recalculate_basler_spot()

            self.lbl_basler_status.setText("狀態: 畫圖成功")
            self.lbl_basler_status.setStyleSheet("color: #2E7D32; font-weight: bold;")
        except Exception as e:
            self.lbl_basler_status.setText("狀態: 讀取失敗")
            self.lbl_basler_status.setStyleSheet("color: #C62828; font-weight: bold;")
            QMessageBox.critical(self, "錯誤", f"處理 Basler 數據/圖像時出錯: {str(e)}")

    def apply_basler_param_file(self):
        if not self.basler_param_path:
            return
        try:
            with open(self.basler_param_path, 'r', encoding='utf-8') as f:
                params = json.load(f)
            
            if "shape_type" in params:
                if params["shape_type"] == "ellipse":
                    self.radio_shape_ellipse.setChecked(True)
                else:
                    self.radio_shape_circle.setChecked(True)

            if "use_threshold" in params:
                self.chk_use_threshold.setChecked(params["use_threshold"])
            if "threshold_percent" in params:
                self.spin_thresh_percent.setValue(params["threshold_percent"])
            if "circle_diameter_px" in params:
                self.spin_circle_diameter.setValue(params["circle_diameter_px"])
            if "ellipse_wx_px" in params:
                self.spin_ellipse_wx.setValue(params["ellipse_wx_px"])
            if "ellipse_wy_px" in params:
                self.spin_ellipse_wy.setValue(params["ellipse_wy_px"])
            if "basler_cross_size" in params:
                self.spin_basler_cross_size.setValue(params["basler_cross_size"])

            if "basler_click_points" in params:
                self.basler_click_points = [tuple(p) for p in params.get("basler_click_points", [])]
                if len(self.basler_click_points) > 0:
                    self.update_basler_measure_display()
                    self.redraw_basler_measure_crosses()

            if "center_x_px" in params and "center_y_px" in params:
                self.basler_center = (params["center_x_px"], params["center_y_px"])
                self.radio_manual_center.setChecked(True)
            
            if "colorbar_levels" in params and self.basler_matrix is not None:
                c_levels = params["colorbar_levels"]
                self.basler_hist.setLevels(c_levels[0], c_levels[1])

            self.update_basler_circle()
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"解析 Basler 參數檔失敗: {str(e)}")

    def recalculate_basler_spot(self):
        if self.basler_matrix is None:
            return
            
        if self.chk_basler_enable_spot.isChecked():
            use_thresh = self.chk_use_threshold.isChecked()
            thresh_percent = self.spin_thresh_percent.value()
            if self.radio_auto_center.isChecked():
                self.basler_center = self._compute_auto_spot_center(
                    self.basler_matrix, "peak_geom", use_thresh, thresh_percent)
            elif self.radio_centroid_center.isChecked():
                self.basler_center = self._compute_auto_spot_center(
                    self.basler_matrix, "centroid", use_thresh, thresh_percent)
            elif self.radio_thresh_geom_center.isChecked():
                self.basler_center = self._compute_auto_spot_center(
                    self.basler_matrix, "thresh_geom", use_thresh, thresh_percent)
            self.update_basler_circle()
        else:
            self.clear_basler_markers()

    def clear_basler_markers(self):
        if self.basler_circle_item is not None:
            self.plot_basler.removeItem(self.basler_circle_item)
            self.basler_circle_item = None
        if self.basler_center_spot is not None:
            self.plot_basler.removeItem(self.basler_center_spot)
            self.basler_center_spot = None
            
        for item in self.basler_fixed_cross_items:
            self.plot_basler.removeItem(item)
        self.basler_fixed_cross_items.clear()

        self.plot_basler_x_profile.clear()
        self.plot_basler_y_profile.clear()
        self.plot_basler_x_log_profile.clear()
        self.plot_basler_y_log_profile.clear()

        self.lbl_b_center.setText("中心座標: (X: --, Y: --)")
        self.lbl_b_peak.setText("最大強度 (Peak): --")
        self.lbl_b_thresh_val.setText("計算門檻值: 未啟用" if not self.chk_use_threshold.isChecked() else "計算門檻值: --")
        self.lbl_b_x_width.setText("X 軸寬度 (@門檻): --")
        self.lbl_b_y_width.setText("Y 軸寬度 (@門檻): --")
        self.lbl_b_dia_px.setText("外框尺寸: -- px")
        self.lbl_b_dia_um.setText("實際尺寸: -- μm")
        self.lbl_b_radius_um.setText("實際半徑: -- μm")
        self.lbl_b_area_um.setText("實際面積: -- μm²")
        self.lbl_b_sum_intensity.setText("總光強度: --")
        self.lbl_b_mean_intensity.setText("平均強度: --")

    def on_basler_colorbar_changed(self):
        if self.basler_matrix is not None:
            min_v, max_v = self.basler_hist.getLevels()
            
            self.plot_basler_x_profile.setYRange(min_v, max_v, padding=0)
            self.plot_basler_y_profile.setYRange(min_v, max_v, padding=0)

            log_min = np.log10(max(min_v, 1e-3))
            log_max = np.log10(max(max_v, 1e-3))
            if log_min < log_max:
                self.plot_basler_x_log_profile.setYRange(log_min, log_max, padding=0)
                self.plot_basler_y_log_profile.setYRange(log_min, log_max, padding=0)

            if self.basler_cross_profile_win is not None and self.basler_cross_profile_win.isVisible():
                cx, cy = self.basler_heatmap_cross_point if self.basler_heatmap_cross_point is not None else (self.basler_matrix.shape[1]//2, self.basler_matrix.shape[0]//2)
                self.basler_cross_profile_win.update_profiles(self.basler_matrix, cx, cy, y_range=(min_v, max_v))

    def update_basler_circle(self):
        if self.basler_matrix is None or not self.chk_basler_enable_spot.isChecked() or self.basler_center is None or self.is_updating_basler_ui:
            if not self.chk_basler_enable_spot.isChecked():
                self.clear_basler_markers()
            return
            
        self.is_updating_basler_ui = True
        self.clear_basler_markers()

        cx, cy = self.basler_center
        h, w = self.basler_matrix.shape

        peak_value = np.max(self.basler_matrix)
        use_thresh = self.chk_use_threshold.isChecked()
        is_ellipse = self.radio_shape_ellipse.isChecked()
        pixel_pitch_um = 3.45

        cy_clamped = max(0, min(h - 1, cy))
        x_profile = self.basler_matrix[cy_clamped, :]
        x_axis = np.arange(len(x_profile))

        cx_clamped = max(0, min(w - 1, cx))
        y_profile = self.basler_matrix[:, cx_clamped]
        y_axis = np.arange(len(y_profile))

        x_width_px, y_width_px = 0, 0
        x_width_um, y_width_um = 0.0, 0.0

        if use_thresh:
            thresh_percent = self.spin_thresh_percent.value()
            thresh_val = peak_value * (thresh_percent / 100.0)

            x_above = np.where(x_profile >= thresh_val)[0]
            if len(x_above) > 1:
                x_width_px = x_above[-1] - x_above[0]
                x_width_um = x_width_px * pixel_pitch_um

            y_above = np.where(y_profile >= thresh_val)[0]
            if len(y_above) > 1:
                y_width_px = y_above[-1] - y_above[0]
                y_width_um = y_width_px * pixel_pitch_um

            if is_ellipse:
                if x_width_px > 0:
                    self.spin_ellipse_wx.blockSignals(True)
                    self.spin_ellipse_wx.setValue(int(x_width_px))
                    self.spin_ellipse_wx.blockSignals(False)
                if y_width_px > 0:
                    self.spin_ellipse_wy.blockSignals(True)
                    self.spin_ellipse_wy.setValue(int(y_width_px))
                    self.spin_ellipse_wy.blockSignals(False)
            else:
                auto_diameter_px = max(x_width_px, y_width_px)
                if auto_diameter_px > 0:
                    self.spin_circle_diameter.blockSignals(True)
                    self.spin_circle_diameter.setValue(int(auto_diameter_px))
                    self.spin_circle_diameter.blockSignals(False)

            self.lbl_b_thresh_val.setText(f"計算門檻值 ({thresh_percent}%): {thresh_val:.1f}")
            self.lbl_b_x_width.setText(f"X 軸寬度 (@門檻): {x_width_px} px ({x_width_um:.2f} μm)")
            self.lbl_b_y_width.setText(f"Y 軸寬度 (@門檻): {y_width_px} px ({y_width_um:.2f} μm)")
        else:
            self.lbl_b_thresh_val.setText("計算門檻值: 未啟用")
            self.lbl_b_x_width.setText("X 軸寬度 (@門檻): --")
            self.lbl_b_y_width.setText("Y 軸寬度 (@門檻): --")

        theta = np.linspace(0, 2*np.pi, 100)

        if is_ellipse:
            wx_px = self.spin_ellipse_wx.value()
            wy_px = self.spin_ellipse_wy.value()
            rx_px = wx_px / 2.0
            ry_px = wy_px / 2.0

            circle_x = cx + rx_px * np.cos(theta)
            circle_y = cy + ry_px * np.sin(theta)

            wx_um = wx_px * pixel_pitch_um
            wy_um = wy_px * pixel_pitch_um
            rx_um = rx_px * pixel_pitch_um
            ry_um = ry_px * pixel_pitch_um
            area_um2 = np.pi * rx_um * ry_um

            y_grid, x_grid = np.ogrid[:h, :w]
            mask = ((x_grid - cx)**2 / (rx_px**2)) + ((y_grid - cy)**2 / (ry_px**2)) <= 1.0

            self.lbl_b_dia_px.setText(f"外框尺寸 (Wx, Wy): {wx_px} × {wy_px} px")
            self.lbl_b_dia_um.setText(f"實際尺寸: {wx_um:.1f} × {wy_um:.1f} μm")
            self.lbl_b_radius_um.setText(f"實際半徑 (Rx, Ry): {rx_um:.1f}, {ry_um:.1f} μm")
        else:
            diameter_px = self.spin_circle_diameter.value()
            radius_px = diameter_px / 2.0

            circle_x = cx + radius_px * np.cos(theta)
            circle_y = cy + radius_px * np.sin(theta)

            diameter_um = diameter_px * pixel_pitch_um
            radius_um = radius_px * pixel_pitch_um
            area_um2 = np.pi * (radius_um ** 2)

            y_grid, x_grid = np.ogrid[:h, :w]
            mask = (x_grid - cx)**2 + (y_grid - cy)**2 <= radius_px**2

            self.lbl_b_dia_px.setText(f"圓圈直徑: {diameter_px} px")
            self.lbl_b_dia_um.setText(f"實際直徑: {diameter_um:.2f} μm")
            self.lbl_b_radius_um.setText(f"實際半徑: {radius_um:.2f} μm")

        v_fixed = pg.PlotCurveItem(x=[cx, cx], y=[0, h], pen=pg.mkPen('y', width=1, style=Qt.DashLine))
        h_fixed = pg.PlotCurveItem(x=[0, w], y=[cy, cy], pen=pg.mkPen('y', width=1, style=Qt.DashLine))
        self.plot_basler.addItem(v_fixed)
        self.plot_basler.addItem(h_fixed)
        self.basler_fixed_cross_items.extend([v_fixed, h_fixed])

        self.basler_center_spot = pg.ScatterPlotItem(x=[cx], y=[cy], symbol='+', size=12, pen=pg.mkPen('r', width=2))
        self.plot_basler.addItem(self.basler_center_spot)
        
        self.basler_circle_item = pg.PlotCurveItem(circle_x, circle_y, pen=pg.mkPen('y', width=2))
        self.plot_basler.addItem(self.basler_circle_item)

        circle_pixels = self.basler_matrix[mask]
        sum_intensity = np.sum(circle_pixels) if len(circle_pixels) > 0 else 0
        mean_intensity = np.mean(circle_pixels) if len(circle_pixels) > 0 else 0

        self.plot_basler_x_profile.clear()
        self.plot_basler_y_profile.clear()
        self.plot_basler_x_log_profile.clear()
        self.plot_basler_y_log_profile.clear()

        curve_x = pg.PlotCurveItem(x_axis, x_profile, pen=pg.mkPen('#00E5FF', width=1.5))
        curve_y = pg.PlotCurveItem(y_axis, y_profile, pen=pg.mkPen('#FF5722', width=1.5))
        self.plot_basler_x_profile.addItem(curve_x)
        self.plot_basler_y_profile.addItem(curve_y)

        x_profile_log = np.log10(np.maximum(x_profile, 1e-3))
        y_profile_log = np.log10(np.maximum(y_profile, 1e-3))

        curve_x_log = pg.PlotCurveItem(x_axis, x_profile_log, pen=pg.mkPen('#00E5FF', width=1.5))
        curve_y_log = pg.PlotCurveItem(y_axis, y_profile_log, pen=pg.mkPen('#FF5722', width=1.5))
        self.plot_basler_x_log_profile.addItem(curve_x_log)
        self.plot_basler_y_log_profile.addItem(curve_y_log)

        if use_thresh:
            line_thresh_x = pg.InfiniteLine(pos=thresh_val, angle=0, pen=pg.mkPen('r', width=1.5, style=Qt.DashLine))
            line_thresh_y = pg.InfiniteLine(pos=thresh_val, angle=0, pen=pg.mkPen('r', width=1.5, style=Qt.DashLine))
            self.plot_basler_x_profile.addItem(line_thresh_x)
            self.plot_basler_y_profile.addItem(line_thresh_y)

            log_thresh_val = np.log10(max(thresh_val, 1e-3))
            line_thresh_x_log = pg.InfiniteLine(pos=log_thresh_val, angle=0, pen=pg.mkPen('r', width=1.5, style=Qt.DashLine))
            line_thresh_y_log = pg.InfiniteLine(pos=log_thresh_val, angle=0, pen=pg.mkPen('r', width=1.5, style=Qt.DashLine))
            self.plot_basler_x_log_profile.addItem(line_thresh_x_log)
            self.plot_basler_y_log_profile.addItem(line_thresh_y_log)

        self.plot_basler_x_profile.setXRange(0, w, padding=0)
        self.plot_basler_y_profile.setXRange(0, h, padding=0)
        self.plot_basler_x_log_profile.setXRange(0, w, padding=0)
        self.plot_basler_y_log_profile.setXRange(0, h, padding=0)

        self.on_basler_colorbar_changed()

        self.lbl_b_center.setText(f"中心座標: (X: {cx}, Y: {cy})")
        self.lbl_b_peak.setText(f"強度 (Peak): {peak_value:.1f}")

        self.lbl_b_area_um.setText(f"實際面積: {area_um2:.2f} μm²")
        self.lbl_b_sum_intensity.setText(f"總光強度: {sum_intensity:.1f}")
        self.lbl_b_mean_intensity.setText(f"平均強度: {mean_intensity:.2f}")

        self.is_updating_basler_ui = False

    def on_shape_type_changed(self):
        is_ellipse = self.radio_shape_ellipse.isChecked()
        self.container_circle_spin.setVisible(not is_ellipse)
        self.container_ellipse_spin.setVisible(is_ellipse)
        self.update_basler_circle()

    def on_threshold_toggled(self, checked):
        self.spin_thresh_percent.setEnabled(checked)
        self.lbl_thresh_spin.setEnabled(checked)
        self.recalculate_basler_spot()

    def on_basler_mouse_moved(self, evt):
        pos = evt
        if self.plot_basler.sceneBoundingRect().contains(pos):
            mouse_point = self.plot_basler.getViewBox().mapSceneToView(pos)
            x = mouse_point.x()
            y = mouse_point.y()
            
            if self.basler_matrix is not None:
                h, w = self.basler_matrix.shape
                if 0 <= x < w and 0 <= y < h:
                    self.basler_v_line.show()
                    self.basler_h_line.show()
                    self.basler_v_line.setPos(x)
                    self.basler_h_line.setPos(y)
                    self.lbl_basler_cursor.setText(f"目前滑鼠: X: {x:.1f}, Y: {y:.1f}")
                else:
                    self.basler_v_line.hide()
                    self.basler_h_line.hide()
                    self.lbl_basler_cursor.setText("目前滑鼠: X: --, Y: --")
            else:
                self.basler_v_line.hide()
                self.basler_h_line.hide()
                self.lbl_basler_cursor.setText("目前滑鼠: X: --, Y: --")
        else:
            self.basler_v_line.hide()
            self.basler_h_line.hide()
            self.lbl_basler_cursor.setText("目前滑鼠: X: --, Y: --")

    def on_basler_scene_clicked(self, evt):
        if self.basler_matrix is None:
            return
            
        pos = evt.scenePos()
        if self.plot_basler.sceneBoundingRect().contains(pos):
            mouse_point = self.plot_basler.getViewBox().mapSceneToView(pos)
            cx = int(round(mouse_point.x()))
            cy = int(round(mouse_point.y()))
            
            h, w = self.basler_matrix.shape
            if 0 <= cx < w and 0 <= cy < h:
                if self.chk_basler_heatmap_cross.isChecked():
                    self.basler_heatmap_cross_point = (cx, cy)
                    self.redraw_basler_heatmap_cross_item()

                    if self.basler_cross_profile_win is not None and self.basler_cross_profile_win.isVisible():
                        min_v, max_v = self.basler_hist.getLevels()
                        self.basler_cross_profile_win.update_profiles(self.basler_matrix, cx, cy, y_range=(min_v, max_v))

                if self.radio_manual_center.isChecked():
                    if self.chk_basler_enable_spot.isChecked():
                        self.basler_center = (cx, cy)
                        self.update_basler_circle()
                else:
                    if self.chk_basler_measure_cross.isChecked():
                        if len(self.basler_click_points) >= 2:
                            self.clear_basler_measure_items_only()
                            self.basler_click_points.clear()
                        self.basler_click_points.append((cx, cy))
                        self.update_basler_measure_display()
                        self.redraw_basler_measure_crosses()

    def redraw_basler_heatmap_cross_item(self):
        for item in self.basler_heatmap_cross_items:
            self.plot_basler.removeItem(item)
        self.basler_heatmap_cross_items.clear()

        if self.basler_heatmap_cross_point is not None and self.chk_basler_heatmap_cross.isChecked():
            cx, cy = self.basler_heatmap_cross_point
            h, w = self.basler_matrix.shape
            pen = pg.mkPen('#00E676', width=2)
            v_item = pg.PlotCurveItem(x=[cx, cx], y=[0, h], pen=pen)
            h_item = pg.PlotCurveItem(x=[0, w], y=[cy, cy], pen=pen)
            self.plot_basler.addItem(v_item)
            self.plot_basler.addItem(h_item)
            self.basler_heatmap_cross_items.extend([v_item, h_item])

    def redraw_basler_measure_crosses(self):
        self.clear_basler_measure_items_only()
        if not self.chk_basler_measure_cross.isChecked():
            return
        half_size = self.spin_basler_cross_size.value() / 2.0
        for idx, (cx, cy) in enumerate(self.basler_click_points):
            color = 'y' if idx == 0 else 'c'
            pen = pg.mkPen(color, width=2)
            v_marker = pg.PlotCurveItem(x=[cx, cx], y=[cy - half_size, cy + half_size], pen=pen)
            h_marker = pg.PlotCurveItem(x=[cx - half_size, cx + half_size], y=[cy, cy], pen=pen)
            self.plot_basler.addItem(v_marker)
            self.plot_basler.addItem(h_marker)
            self.basler_measure_items.extend([v_marker, h_marker])

    def update_basler_measure_display(self):
        pixel_pitch_um = 3.45
        if len(self.basler_click_points) == 1:
            p1 = self.basler_click_points[0]
            self.lbl_basler_distance.setText("位置差距: ΔX: --, ΔY: -- | 總距離: -- px")
            self.lbl_basler_real_distance.setText("實際差距 (* 3.45): ΔX: -- μm, ΔY: -- μm | 總距離: -- μm")
        elif len(self.basler_click_points) == 2:
            p1 = self.basler_click_points[0]
            p2 = self.basler_click_points[1]
            dx = p2[0] - p1[0]
            dy = p2[1] - p1[1]
            distance_px = np.sqrt(dx**2 + dy**2)
            self.lbl_basler_distance.setText(f"位置差距: ΔX: {abs(dx)} px, ΔY: {abs(dy)} px | 總距離: {distance_px:.2f} px")
            
            dx_real = abs(dx) * pixel_pitch_um
            dy_real = abs(dy) * pixel_pitch_um
            distance_real = distance_px * pixel_pitch_um
            self.lbl_basler_real_distance.setText(f"實際差距 (* 3.45): ΔX: {dx_real:.2f} μm, ΔY: {dy_real:.2f} μm | 總距離: {distance_real:.2f} μm")

    def clear_basler_measure_items_only(self):
        for item in self.basler_measure_items:
            self.plot_basler.removeItem(item)
        self.basler_measure_items.clear()

    def clear_basler_measure_points(self):
        self.basler_click_points.clear()
        self.clear_basler_measure_items_only()
        self.lbl_basler_distance.setText("位置差距: ΔX: --, ΔY: -- | 總距離: -- px")
        self.lbl_basler_real_distance.setText("實際差距 (* 3.45): ΔX: -- μm, ΔY: -- μm | 總距離: -- μm")

    def export_basler_results(self):
        if self.basler_matrix is None:
            return
            
        default_name = "Basler_Result"
        if self.basler_img_path:
            default_name = os.path.splitext(os.path.basename(self.basler_img_path))[0]

        initial_path = default_name
        if self.save_dir_path:
            initial_path = os.path.join(self.save_dir_path, default_name)

        path, _ = QFileDialog.getSaveFileName(self, "匯出 Basler 結果", initial_path, "JSON Files (*.json)")
        if not path:
            return

        base_path, ext = os.path.splitext(path)
        if ext.lower() != ".json":
            path = base_path + ".json"

        try:
            self.basler_v_line.hide()
            self.basler_h_line.hide()

            cx, cy = self.basler_center if self.basler_center else (0, 0)
            is_ellipse = self.radio_shape_ellipse.isChecked()
            shape_type = "ellipse" if is_ellipse else "circle"
            dia_px = self.spin_circle_diameter.value()
            wx_px = self.spin_ellipse_wx.value()
            wy_px = self.spin_ellipse_wy.value()

            pixel_pitch_um = 3.45
            
            if is_ellipse:
                rx_px, ry_px = wx_px / 2.0, wy_px / 2.0
                area_um2 = np.pi * (rx_px * pixel_pitch_um) * (ry_px * pixel_pitch_um)
            else:
                radius_um = (dia_px / 2.0) * pixel_pitch_um
                area_um2 = np.pi * (radius_um ** 2)

            peak_value = float(np.max(self.basler_matrix))
            use_thresh = self.chk_use_threshold.isChecked()
            thresh_percent = self.spin_thresh_percent.value()
            thresh_val = peak_value * (thresh_percent / 100.0) if use_thresh else None

            h, w = self.basler_matrix.shape
            x_profile = self.basler_matrix[cy, :]
            y_profile = self.basler_matrix[:, cx]

            if use_thresh and self.basler_center is not None:
                x_above = np.where(x_profile >= thresh_val)[0]
                x_width_px = int(x_above[-1] - x_above[0]) if len(x_above) > 1 else 0
                x_width_um = float(x_width_px * pixel_pitch_um)

                y_above = np.where(y_profile >= thresh_val)[0]
                y_width_px = int(y_above[-1] - y_above[0]) if len(y_above) > 1 else 0
                y_width_um = float(y_width_px * pixel_pitch_um)
            else:
                x_width_px, x_width_um = None, None
                y_width_px, y_width_um = None, None

            if self.basler_center is not None:
                y_grid, x_grid = np.ogrid[:h, :w]
                if is_ellipse:
                    mask = ((x_grid - cx)**2 / ((wx_px/2.0)**2)) + ((y_grid - cy)**2 / ((wy_px/2.0)**2)) <= 1.0
                else:
                    mask = (x_grid - cx)**2 + (y_grid - cy)**2 <= (dia_px/2.0)**2
                
                circle_pixels = self.basler_matrix[mask]
                sum_intensity = float(np.sum(circle_pixels)) if len(circle_pixels) > 0 else 0.0
                mean_intensity = float(np.mean(circle_pixels)) if len(circle_pixels) > 0 else 0.0
            else:
                sum_intensity, mean_intensity = 0.0, 0.0

            c_levels = [float(v) for v in self.basler_hist.getLevels()]

            data = {
                "image_file": os.path.basename(self.basler_img_path),
                "shape_type": shape_type,
                "center_x_px": cx,
                "center_y_px": cy,
                "peak_intensity": peak_value,
                "use_threshold": use_thresh,
                "threshold_percent": thresh_percent,
                "threshold_value": thresh_val,
                "x_width_px_at_thresh": x_width_px,
                "x_width_um_at_thresh": x_width_um,
                "y_width_px_at_thresh": y_width_px,
                "y_width_um_at_thresh": y_width_um,
                "circle_diameter_px": dia_px,
                "ellipse_wx_px": wx_px,
                "ellipse_wy_px": wy_px,
                "pixel_pitch_um": pixel_pitch_um,
                "circle_area_um2": area_um2,
                "sum_intensity": sum_intensity,
                "mean_intensity": mean_intensity,
                "colorbar_levels": c_levels,
                "basler_cross_size": self.spin_basler_cross_size.value(),
                "basler_click_points": self.basler_click_points
            }
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)

            excel_path = f"{base_path}_Spot_Data.xlsx"
            wb = openpyxl.Workbook()
            
            ws_sum = wb.active
            ws_sum.title = "Summary_and_Spot"
            ws_sum.append(["Item", "Value", "Unit"])
            
            summary_rows = [
                ["Image/Data File", os.path.basename(self.basler_img_path), ""],
                ["Shape Type", shape_type, ""],
                ["Center X", cx, "px"],
                ["Center Y", cy, "px"],
                ["Peak Intensity", peak_value, ""],
                ["Use Threshold", "Yes" if use_thresh else "No", ""],
                ["Threshold Percent", thresh_percent, "%"],
                ["Threshold Value", thresh_val if thresh_val else "N/A", ""],
                ["X Width @ Threshold", x_width_px if x_width_px else "N/A", "px"],
                ["X Width @ Threshold (Real)", x_width_um if x_width_um else "N/A", "μm"],
                ["Y Width @ Threshold", y_width_px if y_width_px else "N/A", "px"],
                ["Y Width @ Threshold (Real)", y_width_um if y_width_um else "N/A", "μm"],
                ["Shape Wx / Diameter", wx_px if is_ellipse else dia_px, "px"],
                ["Shape Wy", wy_px if is_ellipse else dia_px, "px"],
                ["Real Area", area_um2, "μm²"],
                ["Sum Intensity", sum_intensity, ""],
                ["Mean Intensity", mean_intensity, ""]
            ]
            for r in summary_rows:
                ws_sum.append(r)

            b_spot_analysis_path = f"{base_path}_Spot_Analysis.xlsx"
            wb_b_spot = openpyxl.Workbook()
            ws_b_spot = wb_b_spot.active
            ws_b_spot.title = "Spot_and_Measurement"
            ws_b_spot.append(["Item", "Value", "Unit"])

            p1_b = self.basler_click_points[0] if len(self.basler_click_points) > 0 else ("--", "--")
            p2_b = self.basler_click_points[1] if len(self.basler_click_points) > 1 else ("--", "--")
            
            dx_b_px, dy_b_px, dist_b_px = "--", "--", "--"
            dx_b_um, dy_b_um, dist_b_um = "--", "--", "--"
            b_pixel_pitch_um = 3.45

            if len(self.basler_click_points) == 2:
                dx_b_px = abs(self.basler_click_points[1][0] - self.basler_click_points[0][0])
                dy_b_px = abs(self.basler_click_points[1][1] - self.basler_click_points[0][1])
                dist_b_px = np.sqrt(dx_b_px**2 + dy_b_px**2)
                dx_b_um = dx_b_px * b_pixel_pitch_um
                dy_b_um = dy_b_px * b_pixel_pitch_um
                dist_b_um = dist_b_px * b_pixel_pitch_um

            b_spot_rows = [
                ["Shape Type", "Ellipse" if is_ellipse else "Circle", ""],
                ["Circle Size / Wx", wx_px if is_ellipse else dia_px, "px"],
                ["Ellipse Wy", wy_px if is_ellipse else "N/A", "px"],
                ["Use Threshold", "Yes" if use_thresh else "No", ""],
                ["Threshold Percent", thresh_percent, "%"],
                ["Mouse Cursor X", "N/A (Realtime)", "px"],
                ["Mouse Cursor Y", "N/A (Realtime)", "px"],
                ["Click Point 1 (X)", p1_b[0] if p1_b != ("--", "--") else "--", "px"],
                ["Click Point 1 (Y)", p1_b[1] if p1_b != ("--", "--") else "--", "px"],
                ["Click Point 2 (X)", p2_b[0] if p2_b != ("--", "--") else "--", "px"],
                ["Click Point 2 (Y)", p2_b[1] if p2_b != ("--", "--") else "--", "px"],
                ["Delta X (px)", dx_b_px, "px"],
                ["Delta Y (px)", dy_b_px, "px"],
                ["Total Distance (px)", dist_b_px, "px"],
                ["Delta X (Real)", dx_b_um, "μm"],
                ["Delta Y (Real)", dy_b_um, "μm"],
                ["Total Distance (Real)", dist_b_um, "μm"],
                ["Cross Marker Size", self.spin_basler_cross_size.value(), "px"]
            ]
            for r in b_spot_rows:
                ws_b_spot.append(r)
            wb_b_spot.save(b_spot_analysis_path)

            ws_p = wb.create_sheet(title="Profiles_Data")
            ws_p.append([
                "X Pos (px)", "X Pos (μm)", "X Intensity", 
                "Y Pos (px)", "Y Pos (μm)", "Y Intensity"
            ])

            max_len = max(len(x_profile), len(y_profile))
            for i in range(max_len):
                row = []
                if i < len(x_profile):
                    row.extend([i, i * pixel_pitch_um, float(x_profile[i])])
                else:
                    row.extend(["", "", ""])
                if i < len(y_profile):
                    row.extend([i, i * pixel_pitch_um, float(y_profile[i])])
                else:
                    row.extend(["", "", ""])
                ws_p.append(row)

            chart_x = LineChart()
            chart_x.title = f"X-Axis Intensity Profile (at Y={cy})"
            chart_x.style = 13
            chart_x.y_axis.title = "Intensity"
            chart_x.x_axis.title = "X Position (px)"
            
            data_x_ref = Reference(ws_p, min_col=3, min_row=1, max_row=len(x_profile)+1)
            cats_x_ref = Reference(ws_p, min_col=1, min_row=2, max_row=len(x_profile)+1)
            chart_x.add_data(data_x_ref, titles_from_data=True)
            chart_x.set_categories(cats_x_ref)
            chart_x.width, chart_x.height = 16, 9
            ws_p.add_chart(chart_x, "H2")

            chart_y = LineChart()
            chart_y.title = f"Y-Axis Intensity Profile (at X={cx})"
            chart_y.style = 13
            chart_y.y_axis.title = "Intensity"
            chart_y.x_axis.title = "Y Position (px)"
            
            data_y_ref = Reference(ws_p, min_col=6, min_row=1, max_row=len(y_profile)+1)
            cats_y_ref = Reference(ws_p, min_col=4, min_row=2, max_row=len(y_profile)+1)
            chart_y.add_data(data_y_ref, titles_from_data=True)
            chart_y.set_categories(cats_y_ref)
            chart_y.width, chart_y.height = 16, 9
            ws_p.add_chart(chart_y, "H18")

            wb.save(excel_path)

            img_path = f"{base_path}_Spot_Map.png"
            exporter = pg_export.ImageExporter(self.plot_basler)
            exporter.export(img_path)

            QMessageBox.information(
                self, 
                "成功", 
                f"Basler 光斑分析與匯出成功！\n\n"
                f"JSON 參數檔: {os.path.basename(path)}\n"
                f"光斑與量測數據 Excel: {os.path.basename(b_spot_analysis_path)}\n"
                f"Excel 檔 (含雙波形與圖表): {os.path.basename(excel_path)}\n"
                f"分析圖檔: {os.path.basename(img_path)}"
            )
        except Exception as e:
            QMessageBox.critical(self, "匯出錯誤", f"匯出 Basler 數據時出錯: {str(e)}")